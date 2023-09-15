import openpyxl
from itertools import islice

# df = openpyxl.load_workbook("csv/RPM_record4.xlsx")
# df = openpyxl.load_workbook("Exp4_beam/RPM_beam4.xlsx")
df = openpyxl.load_workbook("Exp5_beam/RPM_beam5.xlsx")

df1 = df.active

# df_time = openpyxl.load_workbook("csv/time_record4.xlsx")
# df_time = openpyxl.load_workbook("Exp4_beam/time_beam4.xlsx")
df_time = openpyxl.load_workbook("Exp5_beam/time_beam5.xlsx")

df1_time = df_time.active

rpm_1000 = []
ext_1000 = []
new_1000 = []
rpm_2000 = []
ext_2000 = []
new_2000 = []
rpm_3000 = []
ext_3000 = []
new_3000 = []
time1 = []
ext_time1 = []
new_time1 = []

''' Get thrust values '''   
def downsample_to_proportion(rows, proportion=1):
    return list(islice(rows, 0, len(rows), int(1/proportion)))

def getdrillRPM():
  for row in range(2, df1.max_row):
    for col in df1.iter_cols(3,3):
      rpm_1000.append(col[row].value)
    for col in df1.iter_cols(7,7):
      rpm_2000.append(col[row].value)
    for col in df1.iter_cols(11,11):
      rpm_3000.append(col[row].value)
  ext_1000 = downsample_to_proportion(range(0,49), 1.0)
  for i in range(len(ext_1000)):
    new_1000.append(float(rpm_1000[ext_1000[i]]))
  ext_2000 = downsample_to_proportion(range(0,49), 1.0)
  for i in range(len(ext_2000)):
    new_2000.append(float(rpm_2000[ext_2000[i]]))
  ext_3000 = downsample_to_proportion(range(0,49), 1.0)
  for i in range(len(ext_3000)):
    new_3000.append(float(rpm_3000[ext_3000[i]]))
  # print(ext_3000)

  # ext_1000 = downsample_to_proportion(range(0,43), 1)
  # for i in range(len(ext_1000)):
  #   ext_1000.append(rpm_1000[ext_1000[i]])
  # print("1000: ", rpm_1000)
  # print("2000: ", rpm_2000)
  # print("3000: ", rpm_3000)
  # print("ext_1000:", ext_1000)
  # print(len(ext_1000))

  return new_1000, new_2000, new_3000

def getRPM():
  for row in range(2, df1.max_row):
    for col in df1.iter_cols(1,1):
      rpm_1000.append(col[row].value)
    # for col in df1.iter_cols(7,7):
    #   rpm_2000.append(col[row].value)
    # for col in df1.iter_cols(11,11):
    #   rpm_3000.append(col[row].value)
  ext_1000 = downsample_to_proportion(range(0,99), 1.0)
  for i in range(len(ext_1000)):
    new_3000.append(float(rpm_1000[ext_1000[i]]))
  # ext_2000 = downsample_to_proportion(range(0,49), 1.0)
  # for i in range(len(ext_2000)):
  #   new_2000.append(float(rpm_2000[ext_2000[i]]))
  # ext_3000 = downsample_to_proportion(range(0,49), 1.0)
  # for i in range(len(ext_3000)):
  #   new_3000.append(float(rpm_3000[ext_3000[i]]))
  # print(new_3000)
  # print(len(new_3000))
  return new_3000


def getTime():
  for row in range(2, df1.max_row):
    for col in df1_time.iter_cols(1,1):
      time1.append(col[row].value)
    # for col in df1.iter_cols(7,7):
    #   rpm_2000.append(col[row].value)
    # for col in df1.iter_cols(11,11):
    #   rpm_3000.append(col[row].value)
  ext_time1 = downsample_to_proportion(range(0,99), 1.0)
  for i in range(len(ext_time1)):
    new_time1.append(float(time1[ext_time1[i]]))
  # print(time1)
  # print(len(new_time1))
  return new_time1


def main():
  getRPM()
  getTime()


if __name__ == "__main__":
  main()