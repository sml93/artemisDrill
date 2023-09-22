import openpyxl
import numpy as np
import matplotlib.pyplot as plt

from scipy.optimize import curve_fit
from itertools import islice

# df = openpyxl.load_workbook("log0_7sept.xlsx")
# df = openpyxl.load_workbook("log28_11Sept.xlsx")
# df = openpyxl.load_workbook("Exp2/log31_11sept.xlsx")
# df = openpyxl.load_workbook("Exp3/log33_11sept.xlsx")
# df = openpyxl.load_workbook("Exp4_beam/log48_12sept.xlsx")
# df = openpyxl.load_workbook("Exp5_beam/log49_12sept.xlsx")
df = openpyxl.load_workbook("log53_22sept.xlsx")


df1 = df.active

thrust_list1 = []
thrust_list = []
reduced_TList = []
truncate_list = []
reduced_TList2 = []


''' Get thrust values '''   
def downsample_to_proportion(rows, proportion=1):
    return list(islice(rows, 0, len(rows), int(1/proportion)))


def getThrust():
    # for row in range(2, df1.max_row):
    for row in range(2, 893):
        for col in df1.iter_cols(2,2):
            # print(col[row].value)
            thrust_list.append(col[row].value)
    print(len(thrust_list))
    reduced_TList = downsample_to_proportion(range(0,len(thrust_list)), 0.11)
    # reduced_TList2 = downsample_to_proportion(range(0,len(reduced_TList)), 0.3)
    # print(len(reduced_TList))
    for i in range(len(reduced_TList)):
        truncate_list.append(thrust_list[reduced_TList[i]])
    # print((truncate_list))
    return thrust_list, truncate_list



def plotter():    
    # plot_xlist1 = getThrust1()
    plot_xlist, plot_trunc = getThrust()

    # x1data = np.asarray(plot_xlist1)
    xdata = np.asarray(plot_xlist)
    x1data = np.asarray(plot_trunc)
    print(len(x1data))

    # plt.plot(range(len(x1data)), x1data, label="thrust1")
    plt.plot(range(len(xdata)), xdata, label="thrust")
    plt.plot(range(len(x1data)), x1data, label="truncated")
    plt.grid(color='k', linestyle='-', linewidth=0.5)
    plt.legend()
    plt.show()



def main():
    # getThrottle()
    # getThrust()
    # getCurrent()
    # getRPM()
    # getPower()
    plotter()


if __name__ == "__main__":
    main()
    