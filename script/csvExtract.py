import csv
import openpyxl
import numpy as np

from itertools import islice

rpm_list = []
cdist_list = []
trunc_rpm = []
trunc_cdist = []

df = openpyxl.load_workbook("RPM_record.xlsx")
df1 = df.active

df_cdist = openpyxl.load_workbook("cdist_flight1.xlsx")
df_cdist = df_cdist.active


def downsample_to_proportion(rows, proportion=1):
    return list(islice(rows, 0, len(rows), int(1/proportion)))


def RPMextractor():
  for row in range(2, df1.max_row):
    for col in df1.iter_cols(1,1):
      rpm_list.append(float(col[row].value))
  # print(rpm_list)
  trunc_rpm = downsample_to_proportion(range(0,133), 0.3)
  for i in range(len(trunc_rpm)):
    trunc_rpm.append(rpm_list[trunc_rpm[i]])
  print(len(trunc_rpm))
  return trunc_rpm


def cdistExtractor():
  for row in range(2, df_cdist.max_row):
    for col in df_cdist.iter_cols(1,1):
      cdist_list.append(float(col[row].value))
  # print(cdist_list)
  trunc_cdist = downsample_to_proportion(range(0,1466), 0.03)
  for i in range(len(trunc_cdist)):
    trunc_cdist.append(cdist_list[trunc_cdist[i]])
  print(len(trunc_cdist))
  return trunc_cdist
    


RPMextractor()
cdistExtractor()